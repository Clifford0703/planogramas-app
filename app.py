import io
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pdfplumber
import streamlit as st

# --- CONFIGURACIÓN DE PÁGINA STREAMLIT ---
st.set_page_config(
    page_title="PDF Table Extractor",
    page_icon="✨",
    layout="wide",
)

# --- PANEL LATERAL MINIMALISTA ---
st.sidebar.markdown("### ⚙️ Ajustes")
st.sidebar.markdown("---")
st.sidebar.caption("Versión 2.4.0")


# --- INTERFAZ PRINCIPAL ---
st.title("✨ PDF Table Extractor - Planogramas")
st.markdown("Carga un archivo PDF para exportar las tablas directamente a Excel.")
st.markdown("---")


# --- ALGORITMO STRICTO DE EXTRACCIÓN GEOMÉTRICA ---
def extraer_tabla_geometria_estricta(pdf_file):
    datos_procesados = []
    patron_ean = re.compile(r"\b\d{10,14}\b")

    with pdfplumber.open(pdf_file) as pdf:
        for pagina in pdf.pages:
            words = pagina.extract_words(
                x_tolerance=2, y_tolerance=2, keep_blank_chars=False
            )
            if not words:
                continue

            width = pagina.width

            # 1. IDENTIFICAR COORDENADAS X_0 DE LOS ENCABEZADOS
            cabecera_words = [w for w in words if w["top"] < 240]
            encabezados_x = {}

            for w in cabecera_words:
                txt = w["text"].lower().strip()
                x0 = w["x0"]

                if "bandeja" in txt and x0 < 50 and "bandeja" not in encabezados_x:
                    encabezados_x["bandeja"] = x0
                elif txt in ["n°", "n°.", "no", "n"] and "num" not in encabezados_x:
                    encabezados_x["num"] = x0
                elif txt == "ean" and "ean" not in encabezados_x:
                    encabezados_x["ean"] = x0
                elif txt == "nombre" and "nombre" not in encabezados_x:
                    encabezados_x["nombre"] = x0
                elif txt == "marca" and "marca" not in encabezados_x:
                    encabezados_x["marca"] = x0
                elif "desc" in txt and "desc" not in encabezados_x:
                    encabezados_x["desc"] = x0
                elif ("fabri" in txt or "fabricante" in txt) and "fab" not in encabezados_x:
                    encabezados_x["fab"] = x0
                elif "cara" in txt and "caras" not in encabezados_x:
                    encabezados_x["caras"] = x0
                elif "altu" in txt and "alt" not in encabezados_x:
                    encabezados_x["alt"] = x0
                elif "prof" in txt and "prof" not in encabezados_x:
                    encabezados_x["prof"] = x0
                elif "unid" in txt and "unid_band" not in encabezados_x and x0 > (width * 0.70):
                    encabezados_x["unid_band"] = x0
                elif ("tot" in txt or "unidad" in txt) and "tot_unid" not in encabezados_x and x0 > (width * 0.85):
                    encabezados_x["tot_unid"] = x0

            list_x = [
                encabezados_x.get("bandeja", 0.01 * width),
                encabezados_x.get("num", 0.08 * width),
                encabezados_x.get("ean", 0.12 * width),
                encabezados_x.get("nombre", 0.22 * width),
                encabezados_x.get("marca", 0.36 * width),
                encabezados_x.get("desc", 0.46 * width),
                encabezados_x.get("fab", 0.58 * width),
                encabezados_x.get("caras", 0.67 * width),
                encabezados_x.get("alt", 0.73 * width),
                encabezados_x.get("prof", 0.79 * width),
                encabezados_x.get("unid_band", 0.86 * width),
                encabezados_x.get("tot_unid", 0.94 * width),
                width,
            ]

            list_x = sorted(list_x)

            # 2. DEFINIR RANGOS X DE CADA COLUMNA
            limites_columnas = []
            for col_idx in range(12):
                x_start = list_x[col_idx] - 2.0
                x_end = list_x[col_idx + 1] - 2.0 if col_idx < 11 else width + 10.0
                limites_columnas.append((x_start, x_end))

            # 3. IDENTIFICAR FILAS POR BANDEJA / EAN
            lineas_dict = {}
            for w in words:
                y_pos = round(w["top"], 1)
                linea_clave = None
                for y_existente in lineas_dict.keys():
                    if abs(y_existente - y_pos) <= 3.0:
                        linea_clave = y_existente
                        break

                if linea_clave is None:
                    linea_clave = y_pos
                    lineas_dict[linea_clave] = []

                lineas_dict[linea_clave].append(w)

            filas_inicio = []
            y_ordenadas = sorted(lineas_dict.keys())

            for y_pos in y_ordenadas:
                texto_linea = " ".join([w["text"] for w in lineas_dict[y_pos]])
                match_ean = patron_ean.search(texto_linea)
                if match_ean:
                    filas_inicio.append((y_pos, match_ean.group(0)))

            if not filas_inicio:
                continue

            # 4. EXTRACCIÓN POR CELDAS Y RETENCIÓN EXACTA DE VALORES
            for i, (y_inicio, ean_codigo) in enumerate(filas_inicio):
                y_fin = (
                    filas_inicio[i + 1][0]
                    if i + 1 < len(filas_inicio)
                    else y_inicio + 50.0
                )

                words_item = [
                    w
                    for y_k in y_ordenadas
                    if y_inicio - 2.0 <= y_k < y_fin - 2.0
                    for w in lineas_dict[y_k]
                ]

                row_12_cols = [""] * 12

                for c_idx in range(12):
                    x_start, x_end = limites_columnas[c_idx]

                    words_celda = [
                        w for w in words_item if x_start <= w["x0"] < x_end
                    ]

                    if not words_celda:
                        row_12_cols[c_idx] = ""
                        continue

                    renglones_celda = {}
                    for w in words_celda:
                        y_r = round(w["top"], 1)
                        r_clave = None
                        for r_k in renglones_celda.keys():
                            if abs(r_k - y_r) <= 3.0:
                                r_clave = r_k
                                break
                        if r_clave is None:
                            r_clave = y_r
                            renglones_celda[r_clave] = []
                        renglones_celda[r_clave].append(w)

                    lineas_texto = []
                    for y_r in sorted(renglones_celda.keys()):
                        palabras_r = sorted(
                            renglones_celda[y_r], key=lambda x: x["x0"]
                        )
                        texto_r = " ".join([p["text"] for p in palabras_r])
                        if texto_r.strip():
                            lineas_texto.append(texto_r.strip())

                    val_raw = " ".join(lineas_texto).strip()

                    if c_idx >= 7:
                        if "*" in val_raw:
                            row_12_cols[c_idx] = "*"
                        else:
                            nums = re.findall(r"\d+", val_raw)
                            row_12_cols[c_idx] = nums[0] if nums else ""
                    else:
                        row_12_cols[c_idx] = val_raw

                if not row_12_cols[2] or not patron_ean.match(row_12_cols[2]):
                    row_12_cols[2] = ean_codigo

                # RESPALDO DERECHO DE TOTALES
                words_derecha = sorted(
                    [w for w in words_item if w["x0"] > (width * 0.65)],
                    key=lambda x: x["x0"]
                )
                
                valores_derecha = []
                for w in words_derecha:
                    txt = w["text"].strip()
                    if txt.isdigit() or txt == "*":
                        valores_derecha.append(txt)

                if len(valores_derecha) >= 1 and not row_12_cols[11]:
                    row_12_cols[11] = valores_derecha[-1]
                if len(valores_derecha) >= 2 and not row_12_cols[10]:
                    row_12_cols[10] = valores_derecha[-2]
                if len(valores_derecha) >= 3 and not row_12_cols[9]:
                    row_12_cols[9] = valores_derecha[-3]
                if len(valores_derecha) >= 4 and not row_12_cols[8]:
                    row_12_cols[8] = valores_derecha[-4]
                if len(valores_derecha) >= 5 and not row_12_cols[7]:
                    row_12_cols[7] = valores_derecha[-5]

                datos_procesados.append(row_12_cols)

    return datos_procesados


# --- GENERACIÓN DE EXCEL CON FORMATO CORPORATIVO ---
def generar_excel_en_memoria(datos_filas, titulo_categoria):
    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumen y KPIs"
    ws_data = wb.create_sheet(title="Reporte_Implementacion")

    ws_summary.views.sheetView[0].showGridLines = True
    ws_data.views.sheetView[0].showGridLines = True

    font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)

    fill_header = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )
    fill_zebra = PatternFill(
        start_color="F2F5F9", end_color="F2F5F9", fill_type="solid"
    )
    fill_kpi = PatternFill(
        start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    header_border = Border(
        left=Side(style="thin", color="1F497D"),
        right=Side(style="thin", color="1F497D"),
        top=Side(style="thin", color="1F497D"),
        bottom=Side(style="medium", color="1F497D"),
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    headers = [
        "Bandeja",
        "N°",
        "EAN",
        "Nombre",
        "Marca",
        "Desc_A",
        "Fabricante",
        "Caras",
        "Altura",
        "Profundidad",
        "Total Unid en Bandeja",
        "Total_Unidades",
    ]

    ws_data.cell(row=1, column=1, value="METRO HIPER MEJORADO").font = font_title
    ws_data.cell(
        row=2,
        column=1,
        value=f"Reporte de Implementación - Categoría: {titulo_categoria}",
    ).font = font_subtitle

    start_row = 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=start_row, column=col_idx, value=header)
        (
            cell.font,
            cell.fill,
            cell.alignment,
            cell.border,
        ) = (font_header, fill_header, align_center, header_border)

    for r_idx, row_data in enumerate(datos_filas, start_row + 1):
        row_fill = fill_zebra if (r_idx % 2 == 0) else None

        while len(row_data) < 12:
            row_data.append("")

        for c_idx, val in enumerate(row_data[:12], 1):
            cell = ws_data.cell(row=r_idx, column=c_idx)
            cell.font, cell.border = font_body, thin_border
            if row_fill:
                cell.fill = row_fill

            if c_idx in [1, 2]:
                cell.value = str(val) if val != "" else ""
                cell.alignment = align_center
            elif c_idx == 3:  # EAN formato texto
                cell.value = str(val)
                cell.alignment, cell.number_format = align_center, "@"
            elif c_idx in [4, 5, 6, 7]:  # Cadenas de Texto
                cell.value = str(val)
                cell.alignment = align_left
            elif c_idx in [8, 9, 10, 11, 12]:  # Números o Asterisco
                val_str = str(val).strip()
                if val_str == "*":
                    cell.value = "*"
                    cell.alignment = align_center
                elif val_str.isdigit():
                    cell.value = int(val_str)
                    cell.alignment, cell.number_format = align_right, "#,##0"
                else:
                    cell.value = val_str
                    cell.alignment = align_center

    tot_row = len(datos_filas) + start_row + 1
    ws_data.cell(row=tot_row, column=4, value="TOTAL GENERAL").font = font_bold
    ws_data.cell(row=tot_row, column=8, value=f"=SUM(H5:H{tot_row-1})").font = (
        font_bold
    )
    ws_data.cell(row=tot_row, column=11, value=f"=SUM(K5:K{tot_row-1})").font = (
        font_bold
    )

    for col_i in range(1, len(headers) + 1):
        c = ws_data.cell(row=tot_row, column=col_i)
        c.border = Border(
            top=Side(style="thin", color="1F497D"),
            bottom=Side(style="double", color="1F497D"),
        )
        if col_i in [8, 11]:
            c.number_format, c.alignment = "#,##0", align_right

    for col in ws_data.columns:
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = 16

    ws_data.column_dimensions["D"].width = 45
    ws_data.column_dimensions["G"].width = 35
    ws_data.freeze_panes = "A5"

    # RESUMEN Y KPIS
    ws_summary.cell(
        row=1,
        column=1,
        value=f"DASHBOARD CATEGORÍA {titulo_categoria.upper()}",
    ).font = font_title
    ws_summary.cell(
        row=2, column=1, value="Resumen Métrico y Participación"
    ).font = font_subtitle

    kpis = [
        (
            "Total SKUs Registrados",
            f"=COUNTA(Reporte_Implementacion!C5:C{tot_row-1})",
            "B4",
            "C4",
            "B5",
            "C5",
        ),
        (
            "Total Caras Exhibidas",
            f"=SUM(Reporte_Implementacion!H5:H{tot_row-1})",
            "E4",
            "F4",
            "E5",
            "F5",
        ),
        (
            "Total Unidades en Bandeja",
            f"=SUM(Reporte_Implementacion!K5:K{tot_row-1})",
            "H4",
            "I4",
            "H5",
            "I5",
        ),
    ]

    for label, formula, top_l, top_r, bot_l, bot_r in kpis:
        ws_summary.merge_cells(f"{top_l}:{top_r}")
        ws_summary.merge_cells(f"{bot_l}:{bot_r}")

        c_lbl = ws_summary[top_l]
        (
            c_lbl.value,
            c_lbl.font,
            c_lbl.alignment,
            c_lbl.fill,
        ) = (
            label,
            Font(name="Calibri", size=10, bold=True, color="595959"),
            align_center,
            fill_kpi,
        )

        c_val = ws_summary[bot_l]
        (
            c_val.value,
            c_val.font,
            c_val.alignment,
            c_val.fill,
            c_val.number_format,
        ) = (
            formula,
            Font(name="Calibri", size=16, bold=True, color="1F497D"),
            align_center,
            fill_kpi,
            "#,##0",
        )

    # Resumen por Marca
    ws_summary.cell(row=8, column=2, value="Resumen por Marca").font = Font(
        name="Calibri", size=12, bold=True, color="1F497D"
    )
    s2_headers = ["Marca", "Caras Total", "Unidades en Bandeja", "% Caras"]
    for col_idx, h in enumerate(s2_headers, 2):
        c = ws_summary.cell(row=9, column=col_idx, value=h)
        c.font, c.fill, c.alignment = font_header, fill_header, align_center

    marcas_set = list(
        dict.fromkeys(
            [
                r[4]
                for r in datos_filas
                if len(r) > 4 and str(r[4]).strip() != ""
            ]
        )
    )

    for idx, m_name in enumerate(marcas_set, 10):
        ws_summary.cell(row=idx, column=2, value=m_name).font = font_body
        ws_summary.cell(
            row=idx,
            column=3,
            value=f"=SUMIF(Reporte_Implementacion!E$5:E${tot_row-1}, B{idx}, Reporte_Implementacion!H$5:H${tot_row-1})",
        ).font = font_body
        ws_summary.cell(
            row=idx,
            column=4,
            value=f"=SUMIF(Reporte_Implementacion!E$5:E${tot_row-1}, B{idx}, Reporte_Implementacion!K$5:K${tot_row-1})",
        ).font = font_body
        ws_summary.cell(
            row=idx, column=5, value=f"=C{idx}/C${10+len(marcas_set)}"
        ).font = font_body

        ws_summary.cell(row=idx, column=2).alignment = align_left
        ws_summary.cell(row=idx, column=3).alignment = align_right
        ws_summary.cell(row=idx, column=4).alignment = align_right
        ws_summary.cell(row=idx, column=5).alignment = align_right

        ws_summary.cell(row=idx, column=3).number_format = "#,##0"
        ws_summary.cell(row=idx, column=4).number_format = "#,##0"
        ws_summary.cell(row=idx, column=5).number_format = "0.00%"

    tot_m_row = 10 + len(marcas_set)
    ws_summary.cell(row=tot_m_row, column=2, value="Total").font = font_bold
    ws_summary.cell(
        row=tot_m_row, column=3, value=f"=SUM(C10:C{tot_m_row-1})"
    ).font = font_bold
    ws_summary.cell(
        row=tot_m_row, column=4, value=f"=SUM(D10:D{tot_m_row-1})"
    ).font = font_bold
    ws_summary.cell(
        row=tot_m_row, column=5, value=f"=SUM(E10:E{tot_m_row-1})"
    ).font = font_bold

    for c_i in range(2, 6):
        cell = ws_summary.cell(row=tot_m_row, column=c_i)
        cell.border = Border(
            top=Side(style="thin", color="1F497D"),
            bottom=Side(style="double", color="1F497D"),
        )
        if c_i in [3, 4]:
            cell.number_format, cell.alignment = "#,##0", align_right
        elif c_i == 5:
            cell.number_format, cell.alignment = "0.00%", align_right

    ws_summary.column_dimensions["B"].width = 28
    ws_summary.column_dimensions["C"].width = 16
    ws_summary.column_dimensions["D"].width = 22
    ws_summary.column_dimensions["E"].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- SECCIÓN DE CARGA Y PROCESAMIENTO ---
col1, col2 = st.columns([2, 1])

with col1:
    uploaded_file = st.file_uploader("Seleccionar archivo PDF", type=["pdf"])

with col2:
    categoria = st.text_input("Categoría", value="General")

if uploaded_file is not None:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚀 Procesar y Convertir", use_container_width=True):
        with st.spinner("Procesando..."):
            datos = extraer_tabla_geometria_estricta(uploaded_file)

            if datos:
                st.success(f"Se procesaron con éxito **{len(datos)} filas**.")

                excel_bytes = generar_excel_en_memoria(datos, categoria)

                st.download_button(
                    label="📥 Descargar Excel",
                    data=excel_bytes,
                    file_name=f"Reporte_{categoria}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.error("No se encontraron registros en el PDF.")

# --- PIE DE PÁGINA DISCRETO ---
st.markdown("<br><br><hr>", unsafe_allow_html=True)
st.markdown(
    "<div style='text-align: center; color: #999999; font-size: 0.8rem;'>"
    "Desarrollado por Alfredo HM"
    "</div>",
    unsafe_allow_html=True,
)
